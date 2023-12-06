from dataclasses import dataclass

from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from banker.data.category import Category, PaymentType
from banker.writer.interfaces.categories_writer import ICategoriesWriter
from banker.formatter.payment_type_formatter import format_payment_type

from openpyxl import Workbook
from openpyxl.styles import Font


@dataclass(frozen=True)
class Location:
    row: int
    col: int


class ExcelCategoriesWriter(ICategoriesWriter):
    def __init__(self):
        self.__bold_font = Font(bold=True)
        self.__title_cell_location = Location(row=2, col=2)
        self.__table_headers_locations = self.__generate_table_headers_locations()
        self.__next_category_location = self.__make_next_category_location()

    def __make_next_category_location(self):
        return {payment_type: Location(location.row + 1, location.col) for
                payment_type, location in self.__table_headers_locations.items()}

    def __generate_table_headers_locations(self):
        result = {}
        row = self.__title_cell_location.row + 3
        col = self.__title_cell_location.col
        for payment_type in PaymentType:
            result[payment_type] = Location(row=row, col=col)
            col += 2
        return result

    def __set_title(self, sheet: Worksheet, title: str | None):
        if title is not None:
            title_cell: Cell = sheet.cell(row=self.__title_cell_location.row,
                                          column=self.__title_cell_location.col,
                                          value=title)
            title_cell.font = self.__bold_font

    def __set_table_headers(self, sheet: Worksheet):
        for payment_type in PaymentType:
            location = self.__table_headers_locations[payment_type]
            header_cell: Cell = sheet.cell(row=location.row - 1, column=location.col,
                                           value=format_payment_type(payment_type))
            header_cell.font = self.__bold_font
            subheader_value_cell: Cell = sheet.cell(row=location.row, column=location.col, value="Kwota")
            subheader_value_cell.font = self.__bold_font
            subheader_category_cell: Cell = sheet.cell(row=location.row, column=location.col + 1, value="Kategoria")
            subheader_category_cell.font = self.__bold_font

    def __set_categories(self, sheet: Worksheet, categories: list[Category]):
        self.__next_category_location = self.__make_next_category_location()
        for category in categories:
            location = self.__next_category_location[category.get_payment_type()]
            sheet.cell(row=location.row, column=location.col, value=abs(category.value.amount))
            sheet.cell(row=location.row, column=location.col + 1, value=category.get_name())
            self.__next_category_location[category.get_payment_type()] = Location(row=location.row + 1,
                                                                                  col=location.col)

    def __set_tables_formatting(self, sheet: Worksheet):
        for payment_type in PaymentType:
            table_start_location = self.__table_headers_locations[payment_type]
            table_start_string_range = f"{get_column_letter(table_start_location.col)}{table_start_location.row}"
            table_end_location = self.__next_category_location[payment_type]
            table_end_string_range = f"{get_column_letter(table_end_location.col + 1)}{table_end_location.row}"
            table = Table(displayName=payment_type.name,
                          ref=f"{table_start_string_range}:{table_end_string_range}")
            sheet.add_table(table)

    def write_categories(self, categories: list[Category], output_filepath: str, title: str | None = None) -> None:
        workbook = Workbook()
        sheet: Worksheet = workbook.active
        self.__set_title(sheet, title)
        self.__set_table_headers(sheet)
        categories = sorted(categories, key=lambda category: category.value)
        self.__set_categories(sheet, categories)
        self.__set_tables_formatting(sheet)

        workbook.save(output_filepath)
